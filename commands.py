"""Helper Module defining command & utility functions

This module defines each function called on when the user enters any
one of the valid commands, as determined by the main script, as well as
a few utility functions.

It also contains a few string constants like 'PREFIX_IN' to use with any 
messages or prompts.

Attributes:
	PREFIX_IN: a string constant for input prompt
	PREFIX_MSG: a string constant for returning msgs from script
	PREFIX_ERROR: a string constant for displaying important error msgs

Functions:
	pack: Returns a single dict used to store data in json
	unpack: Returns a useable dict, populated with objects
	addItem: Prompts user to add a new auction item to 'items' dict
	remItem: Removes entered item from 'items' dict
	addBid: Prompts user to add a new auction bidder to 'bidder' dict
	remBid: Removes entered bidder from 'bidder' dict
	eItem: Not yet implemented
	eBid: Not yet implemented
	sDict: Saves 'items' AND/OR 'bidders' dicts to persistent file
	lDict: Loads 'items' AND/OR 'bidders' dicts from persistent file
	cDict: Clears 'items' AND/OR 'bidders' dicts of all data in memory
	dItem: Not yet implemented
	dBid: Not yet implemented
	pItem: Not yet implemented
	pBid: Not yet implemented
	rec: Prints a reciept for the entered bidder #
	sold: Records item bought by a bidder, along with purchased price
	loadXL: Loads 'items' data from entered XL file
	exportXL: Exports bidder and item information into a local XL file
	totalAmt: Displays total profit of auction

ToDo:
""" 

import openpyxl as xl
from escpos import escpos, printer
import json
from datetime import date
from itemclass import Item
from bidderclass import Bidder


# Input prefix constants
PREFIX_IN = ")> "
PREFIX_MSG = ")* "
PREFIX_ERROR = "(!) ERROR (!)"


# Try initializing printer object & set a null object to disable
# all printer functions if printer is not properly initialized
p = None

try:
	p = printer.Usb(0x04b8, 0x0e28)
	p.text("\n\n\nInitialized\n\n")
	p.cut()

except:
	print(PREFIX_ERROR, """
USB Printer Not Found!
	--Continuing execution, printer functions are disabled--

	Printer is either disconnected or unique id is not defined.
	Restart program with printer connected to resume normal operation.
		""")


def canUse(usr_in, val):
	"""A function used throughout to ensure that when the user inputs
	a command, their input can be used as the desired type and converts
	their input according to the set parameter type of 'val'"""

	try:
		return val(usr_in)

	except ValueError:
		print(PREFIX_MSG + "Error: Invalid Command Value")
		return False


def pack(dict_in):
	"""Returns a single dict used to store data in json

	Arguments:
		- dict_in (dictionary)

	Converts dict_in, a dictionary full of objects, into a dictionary
	full of lists. Uses the instance method exp() to get a list
	representation of all its fields so that it can be stored as a
	dictionary of lists in json."""

	if len(dict_in) == 0:
		return dict_in

	elif len(dict_in) >= 1:
		packed_dict = {}

		for k, obj in dict_in.items():
			packed_dict[k] = obj.exp()

		return packed_dict



def unpack(dict_in, obj):
	"""Returns a useable dict, populated with objects

	Arguments:
		- dict_in (dictionary)
		- obj (custom object)

	Converts dict_in, a dictionary of lists, into a dictionary of
	objects. Takes obj, either an 'Item' or 'Bidder' object and fills
	'unpacked_dict' with objects whose fields are automatically
	re-populated with data from each nested list and asigned a
	numerical id, in string form, stored as its dictionary key."""
	
	if len(dict_in) == 0:
		return dict_in

	elif len(dict_in) >= 1:
		unpacked_dict = {}

		for k, list_obj in dict_in.items():
			unpacked_dict[k] = obj(data_in=list_obj)

		return unpacked_dict


def addItem(items):
	""""""

	print(PREFIX_MSG + "Not yet implemented")
	pass


def remItem(items, usr_args):
	"""Remove item entry and data from main 'items' dict

	Arguments:
		- items (dictionary)
		- usr_args (list)

	First checks if the value the user entered is valid as a number,
	then checks if the entered value is actually a real id # assigned
	to an item. If both conditions are met, then the 'item' is removed
	from the main 'items' dictionary"""

	usr_in = canUse(usr_args[1], val=int)
	if not usr_in:
		return

	try:
		del items[str(usr_in)]

	except KeyError:
		print(PREFIX_MSG + "Error: Invalid item #")
		return

	print(PREFIX_MSG + "Item entry deleted")


def addBid(bidders):
	"""Creates a new bidder entry

	Arguments:
		- bidders (dictionary)

	Asks the user for further input and checks if the value entered
	is useable as an id #. If so, it then asks for a first & last name
	and then initializes a 'Bidder' object and adds it to the main
	'bidders' dict with the id # as its dictionary key"""

	print("\nEnter bidder #:")
	entry_num = input(PREFIX_IN)

	bidder_id = canUse(entry_num, val=int)
	if not bidder_id:
		return

	print("\nEnter First & Last Name:")
	name = input(PREFIX_IN)

	# Create nested dict
	bidders[entry_num] = Bidder()
	bidders[entry_num].id = bidder_id
	bidders[entry_num].name = name

	return


def remBid(bidders, usr_args):
	"""Removes entered bidder from 'bidder' dict

	Arguments:
		- bidders (dictionary)
		- usr_args (list)

	First checks if the entered value is useable as a bidder id # and
	then checks if the bidder # is actually assigned to a bidder. If so,
	the 'bidder' is removed from the main 'bidders' dictionary"""

	usr_in = canUse(usr_args[1], val=int)
	if not usr_in:
		return

	try:
		del bidders[str(usr_in)]

	except KeyError:
		print(PREFIX_MSG + "Error: Invalid bidder #")
		return

	print(PREFIX_MSG + "Bidder entry deleted.")


def eItem():
	print(PREFIX_MSG + "Not yet implemented")
	pass


def eBid():
	print(PREFIX_MSG + "Not yet implemented")
	pass


def sDict(items, bidders, usr_args):
	"""Saves 'items' AND/OR 'bidders' dicts to persistent file

	Arguments:
		- items (dictionary)
		- bidders (dictionary)
		- usr_args (list)

	Packs 'items' and/or 'bidders' dictionaries so they can be stored
	as dictionaries of lists, which is then useable for json. Then
	stores the data in their respective files"""

	# If user added 'i' option, save items dict into json
	if "i" in usr_args:
		print(PREFIX_MSG + "Saving Items...")
		with open("items.txt", "w", encoding="utf-8") as wf:
			json.dump(pack(items), wf, indent=4)

		print(PREFIX_MSG + "Finished")

	# If user added 'b' option, save bidders dict into json
	if "b" in usr_args:
		print(PREFIX_MSG + "Saving Bidders...")
		with open("bidders.txt", "w", encoding="utf-8") as wf:
			json.dump(pack(bidders), wf, indent=4)

		print(PREFIX_MSG + "Finished")

	return


def lDict(items, bidders, usr_args):
	"""Loads 'items' AND/OR 'bidders' dicts from persistent file

	Arguments:
		- items (dictionary)
		- bidders (dictionary)
		- usr_args (list)
	
	Unpacks 'items' and/or 'bidders' dictionaries from their respective
	json files and converts them back into dictionaries of objects so
	they can be easily used by the program."""

	# If user added 'i' option, load items dict from json
	if "i" in usr_args:
		print(PREFIX_MSG + "Loading items...")
		with open("items.txt", "r", encoding="utf-8") as rf:
			items = unpack(json.load(rf), Item)

		print(PREFIX_MSG + "Finished")

	# If user added 'b' option, load bidders dict from json
	if "b" in usr_args:
		print(PREFIX_MSG + "Loading bidders...")
		with open("bidders.txt", "r", encoding="utf-8") as rf:
			bidders = unpack(json.load(rf), Bidder)

		print(PREFIX_MSG + "Finished")

	return
			

def cDict(items, bidders, usr_args):
	"""Clears 'items' AND/OR 'bidders' dicts of all data in memory

	Arguments:
		- items (dictionary)
		- bidders (dictionary)
		- usr_args (list)

	Clears 'items' and/or 'bidders' dictionaries of all objects from
	memory, but does NOT tamper with persistent item and bidder data
	stored in json files. 

	If the dictionaries in memory are cleared, and the program exited
	with the dictionaries still empty or modified, the new data in
	memory will overwrite the existing persistent data in the respective
	json files."""

	# If user added 'i' option, empty the items dict
	if "i" in usr_args:
		print(PREFIX_MSG + "Clearing Items...")
		items.clear()
		print(PREFIX_MSG + "Finished")

	# If user added 'b' option, empty the bidders dict
	if "b" in usr_args:
		print(PREFIX_MSG + "Clearing Bidders...")
		bidders.clear()
		print(PREFIX_MSG + "Finished")

	return


def dItem():
	print(PREFIX_MSG + "Not yet implemented")
	pass


def dBid():
	print(PREFIX_MSG + "Not yet implemented")
	pass


def pItem():
	print(PREFIX_MSG + "Not yet implemented")
	pass


def pBid():
	print(PREFIX_MSG + "Not yet implemented")
	pass


def rec(items, bidders, usr_args):
	"""Prints a reciept for the entered bidder #

	Arguments:
		- items (dictionary)
		- bidders (dictionary)
		- usr_args (list)"""

	bidder_id = usr_args[1]

	# Check if printer has been properly initialized & if entered bidder
	# is valid
	if p == None:
		print(PREFIX_MSG + "Printer functions disabled.")
		return

	elif bidder_id not in bidders:
		print(PREFIX_MSG + "Error: Invalid bidder ID#")
		return

	# Total up price of each item bought by bidder, also the est-value
	total_price = 0.0
	total_val = 0.0
	for key in bidders[bidder_id].cart:
		total_price += items[key].price
		total_val += items[key].est_val

	# Series of text strings to create receipt format
	p.set(align="center", width=2, height=2)
	p.text("--Mission Hill Auction--")

	p.set(width=1, height=1)
	p.text("\n\nBidder #: " + bidder_id)
	p.text("\n" + bidders[bidder_id].name)

	p.text("\n\nItems\n----------")
	for key in bidders[bidder_id].cart:
		tmp_item = items[key]
		p.text("\n\n#{:<4} {:<16} {:>20.2f}".format(key, tmp_item.desc,
			tmp_item.price))

	p.text("\n\n-{:<20} {:>20.2f}".format("Total", total_price))
	p.text("\n\n\nTax Credit: {:.2f}\n---------------\n{}".format(total_price
		- total_val, date.today()))
	p.cut()

	return


def sold(items, bidders, usr_args):
	"""Add item to bidder 'cart' and set the bid 'price' of the item

	Arguments:
		- items (dictionary)
		- bidders (dictionary)
		- user_args (list)"""
	
	item_id = usr_args[1]
	bidder_id = usr_args[2]
	sold_price = canUse(usr_args[3], float)

	if not sold_price:
 		return

	elif item_id not in items or bidder_id not in bidders:
		print(PREFIX_MSG + "Error: Invalid item or bidder ID#")
		return

	else:
		items[item_id].price = sold_price
		bidders[bidder_id].cart.append(str(items[item_id].id))


def loadXL(items):
	"""Populate 'items' data structure from entered xl file.

	Arguments:
		- items (dictionary)

	Asks the user for input describing the exact filepath and worksheet
	name of the xl file. Iterates over each cell in a set number
	of rows and populates the main 'items' dictionary with 'Item'
	objects, filling each field with its respective data value.

	Note: Throughout the program, but most importantly here, the item &
	bidder id #'s are stored as strings."""

	print("\nEnter XL Sheet file path: (does not need extension .xlsx)")
	fp = input(PREFIX_IN)

	print("\nEnter Worksheet Name:")
	ws_name = input(PREFIX_IN)

	try:
		wb = xl.load_workbook(fp + ".xlsx")
		ws = wb[ws_name]

	except (xl.utils.exceptions.InvalidFileException, KeyError):
		print(PREFIX_MSG + "Error: Invalid Filename path or worksheet")
		return


	# Iterate thru each row and create a nested dict for each entry
	print(PREFIX_MSG + "Loading items from XL...")
	for row in ws.iter_rows(min_row=2, max_row=200,
							max_col=5, values_only=True):
		posid = None

		if row[0] == None:
			posid = "empty"
		else:
			posid = str(int(row[0]))

		if row[1] is None:
			desc = ""
		else:
			desc = row[1]


		if row[2] is None:
			value = 0.01
		else:
			value = row[2]


		# NOTE: change 'or' condition to 'and' to fix empty
		# donator entries in item data
		if row[3] == None or row[4] == None:
			donator = ""
		else:
			donator = str(row[3]) + " " + str(row[4])


		items[posid] = Item()
		items[posid].id = posid
		items[posid].desc = desc
		items[posid].donator = donator
		items[posid].est_val = value

	print(PREFIX_MSG + "Finished")
	return

# TODO
# improve readability/documentation
# implement reuse with autofilled sheets with matching dates
# cleanup code
def exportXL(items, bidders):
	"""Exports bidder and item information into a local XL file"""

	wb = xl.load_workbook("auction-results.xlsx")
	ws_bidders = wb["Bidders-2024"]
	ws_items = wb["Items-2024"]

	item_count = len(items) - 1
	item_pos = 1

	for row in ws_items.iter_rows(min_row=2, max_row=item_count+1, 
								  max_col=6):
		row[0].value = items[str(item_pos)].id
		row[1].value = items[str(item_pos)].desc
		row[2].value = items[str(item_pos)].est_val
		row[3].value = items[str(item_pos)].price
		row[4].value = items[str(item_pos)].donator

		item_pos += 1


	bidder_count = len(bidders) + 1
	bidder_rows = ws_bidders[2:bidder_count]
	bidder_pos = 0

	for key, obj in bidders.items():
		items_bought = ""

		bidder_rows[bidder_pos][0].value = obj.id
		bidder_rows[bidder_pos][1].value = obj.name

		for item in obj.cart:
			items_bought += ", " + item

		bidder_rows[bidder_pos][2].value = items_bought[2:]

		bidder_pos += 1


	wb.save("auction-results.xlsx")



def totalAmt(items, bidders):
	pass
